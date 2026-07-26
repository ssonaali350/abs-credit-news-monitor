#!/usr/bin/env python3
"""One-time export: reads the Position Monitor + Scenario tables from
ABS_Monitor.xlsx and writes a static position_monitor.json for the
dashboard's Holdings table and Portfolio KPI cards.

THIS IS SAMPLE/DEMO DATA (8 illustrative deals), not real fund positions —
same portfolio used throughout the dashboard's Portfolio Watch section.

Worst-case loss is recomputed for all three scenarios (BASE/STRESS/SEVERE)
by replicating the sheet's own formula:
    -MAX(0, (MonthlyPerf_lossrate * scenario_CDR_mult * 3 - TrancheStruct_creditEnhancement) * JH_Position)
Verified against the sheet's own precomputed BASE-scenario values (which
match the active scenario, currently BASE) before trusting STRESS/SEVERE
outputs — see conversation history for the verification.

Run manually whenever ABS_Monitor.xlsx changes:
    python3 export_position_monitor.py /path/to/ABS_Monitor.xlsx
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
OUTPUT_FILE = BASE_DIR / "position_monitor.json"
DEFAULT_XLSX = "/Users/sonalijena/Downloads/ABS_Monitor (1).xlsx"

WCL_PATTERN = re.compile(r"'Monthly Performance'!\$?Q(\d+)\*\$Q\$4\*3-'Tranche Structure'!\$?J\$?(\d+)")

FLAG_TEXT = {"🔴 RED": "RED", "🔴 BREACH": "RED", "🟡 YELLOW": "YELLOW", "🟢 GREEN": "GREEN"}


def clean_flag(raw):
    if not raw:
        return None
    for k, v in FLAG_TEXT.items():
        if k in raw:
            return v
    return raw.strip()


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb_f = load_workbook(xlsx_path, data_only=False)
    wb_v = load_workbook(xlsx_path, data_only=True)
    pm_f, pm_v = wb_f["Position Monitor"], wb_v["Position Monitor"]
    mp_v, ts_v = wb_v["Monthly Performance"], wb_v["Tranche Structure"]

    # Scenario table: rows 4-6, columns H (name) I (spread add) J (CDR mult) K (B/A mult) L (desc)
    scenarios = {}
    for r in (4, 5, 6):
        name = pm_v.cell(row=r, column=8).value
        if not name:
            continue
        scenarios[name] = {
            "spread_add_bps": pm_v.cell(row=r, column=9).value,
            "cdr_mult": pm_v.cell(row=r, column=10).value,
            "ba_mult": pm_v.cell(row=r, column=11).value,
            "description": pm_v.cell(row=r, column=12).value,
        }

    # Data rows: 9 through the row before "PORTFOLIO TOTALS"
    tranches = []
    r = 9
    while True:
        deal_id = pm_v.cell(row=r, column=1).value
        if not deal_id or "PORTFOLIO TOTALS" in str(deal_id):
            break

        jh_position = pm_v.cell(row=r, column=6).value or 0
        wcl_formula = pm_f.cell(row=r, column=25).value or ""
        m = WCL_PATTERN.search(wcl_formula)
        worst_case = {}
        if m:
            q_val = mp_v.cell(row=int(m.group(1)), column=17).value or 0
            j_val = ts_v.cell(row=int(m.group(2)), column=10).value or 0
            for scen_name, scen in scenarios.items():
                loss = -max(0, (q_val * scen["cdr_mult"] * 3 - j_val) * jh_position)
                worst_case[scen_name] = round(loss, 4)

        tranches.append({
            "deal_id": deal_id,
            "tranche": pm_v.cell(row=r, column=2).value,
            "rating": pm_v.cell(row=r, column=3).value,
            "sector": pm_v.cell(row=r, column=4).value,
            "report_month": pm_v.cell(row=r, column=5).value,
            "jh_position_m": jh_position,
            "current_oc_ratio": pm_v.cell(row=r, column=9).value,
            "oc_gap_pp": pm_v.cell(row=r, column=11).value,
            "mtm_pl_m": pm_v.cell(row=r, column=20).value or 0,
            "worst_case_loss_m": worst_case,
            "oc_status_flag": clean_flag(pm_v.cell(row=r, column=23).value),
            "action_flag": clean_flag(pm_v.cell(row=r, column=26).value),
        })
        r += 1

    total_position = sum(t["jh_position_m"] for t in tranches)
    weighted_oc_gap = (
        sum(t["oc_gap_pp"] * t["jh_position_m"] for t in tranches if t["oc_gap_pp"] is not None) / total_position
        if total_position else 0
    )
    flag_counts = {"RED": 0, "YELLOW": 0, "GREEN": 0}
    for t in tranches:
        if t["action_flag"] in flag_counts:
            flag_counts[t["action_flag"]] += 1

    totals = {
        "total_position_m": round(total_position, 2),
        "weighted_avg_oc_gap_pp": round(weighted_oc_gap, 3),
        "flag_counts": flag_counts,
        "aggregate_mtm_pl_m": round(sum(t["mtm_pl_m"] for t in tranches), 4),
        "worst_case_loss_m": {
            scen_name: round(sum(t["worst_case_loss_m"].get(scen_name, 0) for t in tranches), 4)
            for scen_name in scenarios
        },
    }

    output = {"tranches": tranches, "totals": totals, "scenarios": scenarios}
    OUTPUT_FILE.write_text(json.dumps(output, indent=2))
    print(f"Wrote {len(tranches)} tranches to {OUTPUT_FILE.name}")
    print(json.dumps(totals, indent=2))


if __name__ == "__main__":
    main()
